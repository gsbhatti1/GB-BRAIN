"""
GB-BRAIN — Performance Compare (Phase 9)
==========================================
Compares demo performance vs live performance.
Flags divergence (demo win rate much higher than live = overfitting).
Also compares backtest vs demo vs live.

Usage:
    python runtime/performance_compare.py
    python runtime/performance_compare.py --symbol ETH --days 30
    python runtime/performance_compare.py --export comparison.xlsx
"""

import argparse
import logging
import math
import sqlite3
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv

load_dotenv(ROOT / ".env")

logger = logging.getLogger("gb_brain.performance_compare")

DB_PATH = ROOT / "db" / "gb_brain.db"

# Divergence thresholds
DIVERGENCE_WARN_PCT     = 10.0  # flag if win_rate_delta or pf_delta > 10%
BACKTEST_FIT_THRESHOLD  = 15.0  # flag if demo deviates from backtest by > 15%


# ---------------------------------------------------------------------------
# Stats helpers
# ---------------------------------------------------------------------------
def _safe_div(num: float, den: float, default: float = 0.0) -> float:
    return num / den if den != 0 else default


def _compute_stats_from_rows(rows: list[dict]) -> dict:
    """
    Compute standard performance metrics from a list of trade rows.

    Expected keys in each row: pnl, side, fill_price, signal_price
    pnl may not always be present — we approximate from price data if needed.

    Returns dict with: trades, wins, losses, win_rate, total_pnl,
                       avg_win, avg_loss, profit_factor, expectancy
    """
    if not rows:
        return _empty_stats()

    total_pnl = 0.0
    wins       = []
    losses     = []

    for r in rows:
        pnl = r.get("pnl")
        if pnl is None:
            # Approximate from fill / signal price
            fp = r.get("fill_price")
            sp = r.get("signal_price")
            side = (r.get("side") or "BUY").upper()
            if fp is not None and sp is not None and sp != 0:
                if side == "BUY":
                    pnl = (float(fp) - float(sp)) / float(sp) * 100
                else:
                    pnl = (float(sp) - float(fp)) / float(sp) * 100
            else:
                pnl = 0.0

        pnl = float(pnl)
        total_pnl += pnl
        if pnl > 0:
            wins.append(pnl)
        else:
            losses.append(pnl)

    n        = len(rows)
    n_wins   = len(wins)
    n_losses = len(losses)
    win_rate = _safe_div(n_wins, n) * 100

    avg_win  = _safe_div(sum(wins),   n_wins)  if wins   else 0.0
    avg_loss = _safe_div(sum(losses), n_losses) if losses else 0.0

    gross_profit = sum(wins)
    gross_loss   = abs(sum(losses))
    pf           = _safe_div(gross_profit, gross_loss)

    expectancy   = _safe_div(total_pnl, n)

    return {
        "trades":         n,
        "wins":           n_wins,
        "losses":         n_losses,
        "win_rate":       round(win_rate, 2),
        "total_pnl_pct":  round(total_pnl, 4),
        "avg_win_pct":    round(avg_win,  4),
        "avg_loss_pct":   round(avg_loss, 4),
        "profit_factor":  round(pf, 3),
        "expectancy_pct": round(expectancy, 4),
    }


def _empty_stats() -> dict:
    return {
        "trades":         0,
        "wins":           0,
        "losses":         0,
        "win_rate":       float("nan"),
        "total_pnl_pct":  float("nan"),
        "avg_win_pct":    float("nan"),
        "avg_loss_pct":   float("nan"),
        "profit_factor":  float("nan"),
        "expectancy_pct": float("nan"),
    }


def _nan_str(v) -> str:
    """Format a possibly-NaN float for display."""
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "  N/A"
    return f"{v:>7.2f}"


# ---------------------------------------------------------------------------
# PerformanceCompare
# ---------------------------------------------------------------------------
class PerformanceCompare:
    """
    Three-way performance comparison: Backtest vs Demo vs Live.

    Data sources:
      - Backtest: backtest_results table in SQLite
      - Demo:     live_trades where source='demo'
      - Live:     live_trades where source='live'

    Flags:
      - Divergence:    |demo_win_rate - live_win_rate| > DIVERGENCE_WARN_PCT
      - Backtest fit:  |backtest_win_rate - demo_win_rate| > BACKTEST_FIT_THRESHOLD
    """

    def __init__(
        self,
        symbol: Optional[str] = None,
        days:   int = 30,
    ):
        self.symbol = symbol.upper() if symbol else None
        self.days   = days
        self.cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()

        self._backtest: dict = {}
        self._demo:     dict = {}
        self._live:     dict = {}
        self._symbols_to_compare: list[str] = []

    # ------------------------------------------------------------------
    # Data loading
    # ------------------------------------------------------------------
    def _load_trade_rows(self, source: str) -> list[dict]:
        """
        Load rows from live_trades filtered by source, cutoff, and optional symbol.
        """
        if not DB_PATH.exists():
            return []

        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        try:
            exists = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='live_trades'"
            ).fetchone()
            if not exists:
                return []

            params: list = [source, self.cutoff]
            query = """
                SELECT *
                FROM   live_trades
                WHERE  source = ?
                  AND  run_at >= ?
            """
            if self.symbol:
                query += " AND symbol = ?"
                params.append(self.symbol)
            query += " ORDER BY run_at"
            rows = conn.execute(query, params).fetchall()
            return [dict(r) for r in rows]
        finally:
            conn.close()

    def _load_backtest_rows(self) -> list[dict]:
        """
        Load rows from backtest_results table.
        Falls back gracefully if the table is absent.
        """
        if not DB_PATH.exists():
            return []

        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        try:
            # Try common table names
            for table in ("backtest_results", "backtest_trades", "backtests"):
                exists = conn.execute(
                    f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table}'"
                ).fetchone()
                if exists:
                    params: list = []
                    query = f"SELECT * FROM {table}"
                    filters = []
                    if self.symbol:
                        filters.append("symbol = ?")
                        params.append(self.symbol)
                    if filters:
                        query += " WHERE " + " AND ".join(filters)
                    rows = conn.execute(query, params).fetchall()
                    return [dict(r) for r in rows]
            logger.info("No backtest table found — backtest stats will be N/A")
            return []
        finally:
            conn.close()

    def _load_backtest_stats(self, symbol: Optional[str]) -> dict:
        """Load and aggregate backtest stats for a symbol."""
        rows = self._load_backtest_rows()
        if symbol:
            rows = [r for r in rows if r.get("symbol", "").upper() == symbol.upper()]
        return _compute_stats_from_rows(rows)

    def _load_demo_stats(self, symbol: Optional[str]) -> dict:
        """Load and aggregate demo trade stats for a symbol."""
        rows = self._load_trade_rows("demo")
        if symbol:
            rows = [r for r in rows if r.get("symbol", "").upper() == symbol.upper()]
        return _compute_stats_from_rows(rows)

    def _load_live_stats(self, symbol: Optional[str]) -> dict:
        """Load and aggregate live trade stats for a symbol."""
        rows = self._load_trade_rows("live")
        if symbol:
            rows = [r for r in rows if r.get("symbol", "").upper() == symbol.upper()]
        return _compute_stats_from_rows(rows)

    # ------------------------------------------------------------------
    # Divergence analysis
    # ------------------------------------------------------------------
    def _compute_divergence(self, demo_stats: dict, live_stats: dict) -> dict:
        """
        Compute delta between demo and live performance.
        Flags if win_rate_delta or pf_delta > DIVERGENCE_WARN_PCT.

        Returns dict: win_rate_delta, pf_delta, flagged (bool), flags (list of str)
        """
        demo_wr = demo_stats.get("win_rate", float("nan"))
        live_wr = live_stats.get("win_rate", float("nan"))
        demo_pf = demo_stats.get("profit_factor", float("nan"))
        live_pf = live_stats.get("profit_factor", float("nan"))

        if math.isnan(demo_wr) or math.isnan(live_wr):
            wr_delta = float("nan")
        else:
            wr_delta = demo_wr - live_wr

        if math.isnan(demo_pf) or math.isnan(live_pf):
            pf_delta = float("nan")
        else:
            pf_delta = demo_pf - live_pf

        flags    = []
        flagged  = False

        if not math.isnan(wr_delta) and abs(wr_delta) > DIVERGENCE_WARN_PCT:
            flags.append(
                f"Win rate divergence: demo={demo_wr:.1f}% live={live_wr:.1f}% "
                f"(Δ={wr_delta:+.1f}% > {DIVERGENCE_WARN_PCT}%)"
            )
            flagged = True

        if not math.isnan(pf_delta) and abs(pf_delta) > DIVERGENCE_WARN_PCT / 10:
            flags.append(
                f"Profit factor divergence: demo={demo_pf:.2f} live={live_pf:.2f} "
                f"(Δ={pf_delta:+.2f})"
            )
            flagged = True

        return {
            "win_rate_delta": round(wr_delta, 2) if not math.isnan(wr_delta) else float("nan"),
            "pf_delta":       round(pf_delta, 3) if not math.isnan(pf_delta) else float("nan"),
            "flagged":        flagged,
            "flags":          flags,
        }

    def _compute_backtest_fit(self, backtest_stats: dict, demo_stats: dict) -> dict:
        """
        Check if demo performance matches backtest within BACKTEST_FIT_THRESHOLD%.
        Overfitting warning if backtest is significantly better than demo.

        Returns dict: wr_fit_pct, pf_fit_pct, fitted (bool), notes (list)
        """
        bt_wr  = backtest_stats.get("win_rate", float("nan"))
        d_wr   = demo_stats.get("win_rate", float("nan"))
        bt_pf  = backtest_stats.get("profit_factor", float("nan"))
        d_pf   = demo_stats.get("profit_factor", float("nan"))

        notes  = []
        fitted = True

        if not math.isnan(bt_wr) and not math.isnan(d_wr):
            wr_gap = abs(bt_wr - d_wr)
            if wr_gap > BACKTEST_FIT_THRESHOLD:
                notes.append(
                    f"Win rate gap: backtest={bt_wr:.1f}% demo={d_wr:.1f}% "
                    f"(Δ={wr_gap:.1f}% > {BACKTEST_FIT_THRESHOLD}% threshold)"
                )
                fitted = False
        else:
            wr_gap = float("nan")

        if not math.isnan(bt_pf) and not math.isnan(d_pf):
            pf_gap = abs(bt_pf - d_pf)
            if pf_gap > BACKTEST_FIT_THRESHOLD / 100:
                notes.append(
                    f"Profit factor gap: backtest={bt_pf:.2f} demo={d_pf:.2f} "
                    f"(Δ={pf_gap:.2f})"
                )
        else:
            pf_gap = float("nan")

        return {
            "wr_fit_gap_pct": round(wr_gap, 2) if not math.isnan(wr_gap) else float("nan"),
            "pf_fit_gap":     round(pf_gap, 3) if not math.isnan(pf_gap) else float("nan"),
            "fitted":         fitted,
            "notes":          notes,
        }

    # ------------------------------------------------------------------
    # Main entry
    # ------------------------------------------------------------------
    def run(self):
        """Load all three sources, compute comparison, print report."""
        logger.info(
            "PerformanceCompare.run() — symbol=%s days=%d",
            self.symbol or "ALL", self.days,
        )

        # Determine which symbols to analyse
        if self.symbol:
            symbols = [self.symbol]
        else:
            symbols = self._discover_symbols()

        if not symbols:
            print("No trade data found for any symbol.")
            return

        self._symbols_to_compare = symbols

        # Pre-load all stats
        for sym in symbols:
            self._backtest[sym] = self._load_backtest_stats(sym)
            self._demo[sym]     = self._load_demo_stats(sym)
            self._live[sym]     = self._load_live_stats(sym)

        self.print_comparison_table()

    def _discover_symbols(self) -> list[str]:
        """Find all symbols that have demo or live trades in the window."""
        if not DB_PATH.exists():
            return []
        conn = sqlite3.connect(DB_PATH)
        try:
            exists = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='live_trades'"
            ).fetchone()
            if not exists:
                return []
            rows = conn.execute(
                """
                SELECT DISTINCT symbol
                FROM   live_trades
                WHERE  source IN ('demo', 'live')
                  AND  run_at >= ?
                ORDER  BY symbol
                """,
                (self.cutoff,),
            ).fetchall()
            return [r[0] for r in rows]
        finally:
            conn.close()

    # ------------------------------------------------------------------
    # Comparison table
    # ------------------------------------------------------------------
    def print_comparison_table(self):
        """Print a 3-column table: Backtest | Demo | Live per metric."""
        metrics = [
            ("Trades",          "trades",         "d"),
            ("Win Rate %",      "win_rate",        "f"),
            ("Profit Factor",   "profit_factor",   "f"),
            ("Total PnL %",     "total_pnl_pct",   "f"),
            ("Avg Win %",       "avg_win_pct",     "f"),
            ("Avg Loss %",      "avg_loss_pct",    "f"),
            ("Expectancy %",    "expectancy_pct",  "f"),
        ]

        print(f"\n{'=' * 72}")
        print(f"  GB-BRAIN Performance Comparison  |  Last {self.days} days")
        print(f"{'=' * 72}")

        for sym in self._symbols_to_compare:
            bt = self._backtest.get(sym, _empty_stats())
            dm = self._demo.get(sym,     _empty_stats())
            lv = self._live.get(sym,     _empty_stats())

            print(f"\n  Symbol: {sym}")
            print(f"  {'Metric':<20} {'Backtest':>12} {'Demo':>12} {'Live':>12}")
            print(f"  {'-' * 60}")

            for label, key, fmt in metrics:
                bv = bt.get(key)
                dv = dm.get(key)
                lv_ = lv.get(key)
                if fmt == "d":
                    bs = f"{int(bv):>12}" if bv is not None and not (isinstance(bv, float) and math.isnan(bv)) else "         N/A"
                    ds = f"{int(dv):>12}" if dv is not None and not (isinstance(dv, float) and math.isnan(dv)) else "         N/A"
                    ls = f"{int(lv_):>12}" if lv_ is not None and not (isinstance(lv_, float) and math.isnan(lv_)) else "         N/A"
                else:
                    bs = _nan_str(bv).rjust(12)
                    ds = _nan_str(dv).rjust(12)
                    ls = _nan_str(lv_).rjust(12)
                print(f"  {label:<20} {bs} {ds} {ls}")

            # Divergence analysis
            div  = self._compute_divergence(dm, lv)
            fit  = self._compute_backtest_fit(bt, dm)

            print()
            if div["flagged"]:
                print(f"  ⚠  DIVERGENCE FLAGS:")
                for flag in div["flags"]:
                    print(f"     • {flag}")

            if not fit["fitted"]:
                print(f"  ⚠  BACKTEST FIT FLAGS:")
                for note in fit["notes"]:
                    print(f"     • {note}")

            if not div["flagged"] and fit["fitted"]:
                print(f"  ✓  No significant divergence detected")

        print(f"\n{'=' * 72}\n")

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------
    def export_xlsx(self, path: str):
        """Export the three-way comparison to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.error("openpyxl not installed — run: pip install openpyxl")
            print("ERROR: openpyxl not installed. Run: pip install openpyxl")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Performance Comparison"

        # Pre-load if not already done
        if not self._symbols_to_compare:
            self.run()

        metrics = [
            ("Trades",         "trades"),
            ("Win Rate %",     "win_rate"),
            ("Profit Factor",  "profit_factor"),
            ("Total PnL %",    "total_pnl_pct"),
            ("Avg Win %",      "avg_win_pct"),
            ("Avg Loss %",     "avg_loss_pct"),
            ("Expectancy %",   "expectancy_pct"),
        ]

        header_fill = PatternFill("solid", fgColor="003366")
        header_font = Font(bold=True, color="FFFFFF")
        sym_fill    = PatternFill("solid", fgColor="DDEEFF")

        row_num = 1
        for sym in self._symbols_to_compare:
            bt = self._backtest.get(sym, _empty_stats())
            dm = self._demo.get(sym,     _empty_stats())
            lv = self._live.get(sym,     _empty_stats())

            # Symbol header
            sym_cell = ws.cell(row=row_num, column=1, value=f"Symbol: {sym}")
            sym_cell.fill = sym_fill
            sym_cell.font = Font(bold=True)
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
            row_num += 1

            # Column headers
            for col, label in enumerate(["Metric", "Backtest", "Demo", "Live"], start=1):
                cell = ws.cell(row=row_num, column=col, value=label)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            row_num += 1

            # Data rows
            for label, key in metrics:
                ws.cell(row=row_num, column=1, value=label)
                for col, stats in enumerate([bt, dm, lv], start=2):
                    v = stats.get(key)
                    ws.cell(row=row_num, column=col, value=v if not (isinstance(v, float) and math.isnan(v)) else None)
                row_num += 1

            # Divergence
            div = self._compute_divergence(dm, lv)
            fit = self._compute_backtest_fit(bt, dm)

            if div["flags"] or fit["notes"]:
                row_num += 1
                ws.cell(row=row_num, column=1, value="Flags").font = Font(bold=True, color="FF0000")
                row_num += 1
                for flag in div["flags"] + fit["notes"]:
                    ws.cell(row=row_num, column=1, value=f"! {flag}")
                    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
                    row_num += 1

            row_num += 2  # blank rows between symbols

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        wb.save(path)
        logger.info("Exported performance comparison to %s", path)
        print(f"Exported to {path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="GB-BRAIN Performance Compare — backtest vs demo vs live",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--symbol", help="Filter by symbol (e.g. ETH, US30)")
    p.add_argument("--days",   type=int, default=30, help="Lookback window in days (default 30)")
    p.add_argument("--export", metavar="PATH", help="Export to Excel (.xlsx)")
    p.add_argument("--log-level", default="INFO", choices=["DEBUG", "INFO", "WARNING", "ERROR"])
    return p


def main():
    args = _build_parser().parse_args()
    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    pc = PerformanceCompare(symbol=args.symbol, days=args.days)
    pc.run()

    if args.export:
        pc.export_xlsx(args.export)


if __name__ == "__main__":
    main()
