"""
GB-BRAIN — Fill Tracker
========================
Queries live_trades SQLite table and produces fill quality reports.
Compares signal entry_price vs actual fill_price to measure slippage.
Tracks execution ratio: signals generated vs signals actually executed.

Usage:
    python runtime/fill_tracker.py
    python runtime/fill_tracker.py --broker oanda --days 7
    python runtime/fill_tracker.py --export fills_report.xlsx
"""

import argparse
import logging
import sqlite3
import sys
from datetime import datetime, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv

load_dotenv(ROOT / ".env")

logger = logging.getLogger("gb_brain.fill_tracker")

DB_PATH = ROOT / "db" / "gb_brain.db"


class FillTracker:
    """
    Analyses fill quality from live_trades + observed_signals tables.

    Metrics computed:
      - Execution ratio (fills / signals generated for same period + symbol)
      - Slippage stats: mean, median, max by broker × symbol
      - Fill quality: % of fills within 0.1% of signal price
    """

    QUALITY_THRESHOLD_PCT = 0.1  # fills within this % count as "high quality"

    def __init__(self, broker: str | None = None, days: int = 7):
        self.broker = broker
        self.days   = days
        self.cutoff = (datetime.utcnow() - timedelta(days=days)).isoformat()

        # Cached data frames (list of dicts)
        self._fills   = []
        self._signals = []

    # ------------------------------------------------------------------
    # Data loading
    # ------------------------------------------------------------------
    def _load_fills(self) -> list[dict]:
        """
        SELECT from live_trades where source in ('demo', 'live').
        Optionally filtered by broker and date window.
        """
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        try:
            params: list = ["demo", "live", self.cutoff]
            query = """
                SELECT *
                FROM   live_trades
                WHERE  source IN (?, ?)
                  AND  run_at >= ?
            """
            if self.broker:
                query += " AND broker = ?"
                params.append(self.broker)
            query += " ORDER BY run_at"
            rows = conn.execute(query, params).fetchall()
            return [dict(r) for r in rows]
        finally:
            conn.close()

    def _load_signals(self) -> list[dict]:
        """
        SELECT from observed_signals for same period.
        Falls back gracefully if table doesn't exist.
        """
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        try:
            # Check table exists
            exists = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='observed_signals'"
            ).fetchone()
            if not exists:
                logger.warning("observed_signals table not found — execution ratio will be N/A")
                return []

            params: list = [self.cutoff]
            query = """
                SELECT *
                FROM   observed_signals
                WHERE  observed_at >= ?
            """
            if self.broker:
                query += " AND broker = ?"
                params.append(self.broker)
            query += " ORDER BY observed_at"
            rows = conn.execute(query, params).fetchall()
            return [dict(r) for r in rows]
        finally:
            conn.close()

    # ------------------------------------------------------------------
    # Computation
    # ------------------------------------------------------------------
    def _compute_execution_ratio(self) -> dict[str, float]:
        """
        Returns {symbol: ratio} where ratio = fills / signals.
        Grouped by symbol, for the loaded period.
        """
        from collections import defaultdict

        sig_counts  = defaultdict(int)
        fill_counts = defaultdict(int)

        for s in self._signals:
            sym = s.get("symbol") or s.get("instrument", "UNKNOWN")
            sig_counts[sym] += 1

        for f in self._fills:
            sym = f.get("symbol", "UNKNOWN")
            fill_counts[sym] += 1

        all_symbols = set(sig_counts) | set(fill_counts)
        ratios: dict[str, float] = {}
        for sym in sorted(all_symbols):
            sigs  = sig_counts.get(sym, 0)
            fills = fill_counts.get(sym, 0)
            if sigs == 0:
                ratios[sym] = float("nan")
            else:
                ratios[sym] = round(fills / sigs * 100, 2)
        return ratios

    def _compute_slippage_stats(self) -> list[dict]:
        """
        Returns list of dicts with mean/median/max slippage by broker + symbol.
        """
        from collections import defaultdict
        import statistics

        groups: dict[tuple, list] = defaultdict(list)
        for f in self._fills:
            key = (f.get("broker", "?"), f.get("symbol", "?"))
            slip = f.get("slippage_pct")
            if slip is not None:
                groups[key].append(float(slip))

        results = []
        for (broker, symbol), slips in sorted(groups.items()):
            results.append(
                {
                    "broker":          broker,
                    "symbol":          symbol,
                    "count":           len(slips),
                    "mean_slip_pct":   round(statistics.mean(slips), 4),
                    "median_slip_pct": round(statistics.median(slips), 4),
                    "max_slip_pct":    round(max(slips), 4),
                }
            )
        return results

    def _compute_fill_quality(self) -> dict[tuple, float]:
        """
        Returns {(broker, symbol): pct_within_threshold} where threshold = 0.1%.
        """
        from collections import defaultdict

        totals = defaultdict(int)
        within = defaultdict(int)

        for f in self._fills:
            key  = (f.get("broker", "?"), f.get("symbol", "?"))
            slip = f.get("slippage_pct")
            totals[key] += 1
            if slip is not None and abs(float(slip)) <= self.QUALITY_THRESHOLD_PCT:
                within[key] += 1

        quality = {}
        for key, total in totals.items():
            quality[key] = round(within[key] / total * 100, 2) if total else 0.0
        return quality

    # ------------------------------------------------------------------
    # Main entry
    # ------------------------------------------------------------------
    def run(self):
        """Load data, compute all stats, print report."""
        logger.info("FillTracker loading data — broker=%s days=%d", self.broker, self.days)
        self._fills   = self._load_fills()
        self._signals = self._load_signals()

        if not self._fills:
            print(f"No fills found for broker={self.broker or 'ALL'} in last {self.days} days.")
            return

        self.print_report()

    def print_report(self):
        """Formatted multi-section report."""
        print("\n" + "=" * 65)
        print(f"  GB-BRAIN Fill Quality Report  |  Last {self.days} days")
        if self.broker:
            print(f"  Broker filter: {self.broker.upper()}")
        print("=" * 65)

        # --- Execution ratio ---
        exec_ratios = self._compute_execution_ratio()
        if exec_ratios:
            print("\n[Execution Ratio — Fills / Signals]")
            print(f"  {'Symbol':<12} {'Signals':>8} {'Fills':>8} {'Exec %':>8}")
            print("  " + "-" * 40)
            for sym, ratio in exec_ratios.items():
                sigs  = sum(1 for s in self._signals if (s.get("symbol") or s.get("instrument")) == sym)
                fills = sum(1 for f in self._fills if f.get("symbol") == sym)
                ratio_str = f"{ratio:.1f}%" if ratio == ratio else "N/A"
                print(f"  {sym:<12} {sigs:>8} {fills:>8} {ratio_str:>8}")

        # --- Slippage stats ---
        slip_stats = self._compute_slippage_stats()
        if slip_stats:
            print("\n[Slippage Statistics]")
            print(
                f"  {'Broker':<10} {'Symbol':<12} {'N':>5} "
                f"{'Mean%':>8} {'Median%':>9} {'Max%':>8}"
            )
            print("  " + "-" * 58)
            for s in slip_stats:
                print(
                    f"  {s['broker']:<10} {s['symbol']:<12} {s['count']:>5} "
                    f"{s['mean_slip_pct']:>8.4f} {s['median_slip_pct']:>9.4f} "
                    f"{s['max_slip_pct']:>8.4f}"
                )

        # --- Fill quality ---
        quality = self._compute_fill_quality()
        if quality:
            print(f"\n[Fill Quality — % within {self.QUALITY_THRESHOLD_PCT}% of signal price]")
            print(f"  {'Broker':<10} {'Symbol':<12} {'Quality%':>10}")
            print("  " + "-" * 36)
            for (broker, symbol), pct in sorted(quality.items()):
                flag = "  *** LOW ***" if pct < 80 else ""
                print(f"  {broker:<10} {symbol:<12} {pct:>9.1f}%{flag}")

        print("\n" + "=" * 65 + "\n")

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------
    def export_xlsx(self, path: str):
        """Export fills data and slippage stats to Excel with openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.error("openpyxl not installed — cannot export. Run: pip install openpyxl")
            return

        self._fills   = self._fills   or self._load_fills()
        self._signals = self._signals or self._load_signals()

        wb = openpyxl.Workbook()

        # Sheet 1: Raw fills
        ws_fills = wb.active
        ws_fills.title = "Fills"
        if self._fills:
            headers = list(self._fills[0].keys())
            ws_fills.append(headers)
            for row in self._fills:
                ws_fills.append([row.get(h) for h in headers])
            for cell in ws_fills[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DDEEFF")

        # Sheet 2: Slippage stats
        ws_slip = wb.create_sheet("Slippage Stats")
        slip_stats = self._compute_slippage_stats()
        slip_headers = ["broker", "symbol", "count", "mean_slip_pct", "median_slip_pct", "max_slip_pct"]
        ws_slip.append(slip_headers)
        for row in slip_stats:
            ws_slip.append([row.get(h) for h in slip_headers])
        for cell in ws_slip[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEEFF")

        # Sheet 3: Fill quality
        ws_qual = wb.create_sheet("Fill Quality")
        quality = self._compute_fill_quality()
        ws_qual.append(["broker", "symbol", f"quality_pct (within {self.QUALITY_THRESHOLD_PCT}%)"])
        for (broker, symbol), pct in sorted(quality.items()):
            ws_qual.append([broker, symbol, pct])
        for cell in ws_qual[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEEFF")

        wb.save(path)
        logger.info("Exported fill report to %s", path)
        print(f"Exported to {path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="GB-BRAIN Fill Tracker — fill quality and slippage analysis",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--broker", help="Filter by broker (blofin / oanda)")
    p.add_argument("--days",   type=int, default=7, help="Lookback window in days (default 7)")
    p.add_argument("--export", metavar="PATH", help="Export to Excel (.xlsx)")
    p.add_argument(
        "--log-level",
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
    )
    return p


def main():
    args = _build_parser().parse_args()
    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    tracker = FillTracker(broker=args.broker, days=args.days)
    tracker.run()

    if args.export:
        tracker.export_xlsx(args.export)


if __name__ == "__main__":
    main()
