"""
GB-BRAIN — Score & Sort
=========================
Reads backtest results from SQLite, computes composite scores,
classifies strategies (GEM / PASS / TRASH), and moves files.

Usage:
    python backtest/score_and_sort.py
    python backtest/score_and_sort.py --report    # Also generate Excel
"""

import sys
import json
import shutil
import argparse
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from config.settings import (
    GEM_WIN_RATE, PASS_WIN_RATE, SCORE_WEIGHTS, MIN_TRADES,
    BEST_DIR, ARCHIVE_DIR, STAGING_DIR,
)
from db.brain_db import connect, update_strategy_status


def compute_composite_score(row):
    """Weighted composite score. Higher = better."""
    w = SCORE_WEIGHTS
    score = (
        (row.get("win_rate", 0) or 0) * w["win_rate"]
        + (row.get("total_return", 0) or 0) * w["total_return"]
        + (row.get("profit_factor", 0) or 0) * w["profit_factor"] * 10  # Scale PF
        + min((row.get("total_trades", 0) or 0), 200) / 200 * 100 * w["trade_count"]
    )
    return round(score, 2)


def classify(win_rate, total_return, profit_factor):
    """Classify a strategy based on metrics."""
    if win_rate >= GEM_WIN_RATE and total_return > 0 and profit_factor > 1.0:
        return "GEM"
    if win_rate < PASS_WIN_RATE or total_return < -20:
        return "TRASH"
    return "PASS"


def run_scoring(generate_report=False):
    """Score all backtest results and classify strategies."""
    conn = connect()

    # Get aggregated results per strategy (best result per strategy)
    df = pd.read_sql(
        """
        SELECT
            s.id as strategy_id,
            s.name as strategy_name,
            s.category,
            s.source_file,
            s.indicators,
            br.symbol,
            br.timeframe,
            br.total_trades,
            br.win_rate,
            br.total_return,
            br.max_drawdown,
            br.profit_factor,
            br.sharpe_ratio
        FROM backtest_results br
        JOIN strategies s ON s.id = br.strategy_id
        WHERE br.total_trades >= ?
        ORDER BY br.win_rate DESC
        """,
        conn,
        params=(MIN_TRADES,),
    )

    if df.empty:
        print("[INFO] No backtest results to score.")
        conn.close()
        return

    # Compute scores
    df["composite_score"] = df.apply(compute_composite_score, axis=1)
    df["status"] = df.apply(
        lambda r: classify(r["win_rate"], r["total_return"], r["profit_factor"]),
        axis=1,
    )
    df = df.sort_values("composite_score", ascending=False)

    # Update database
    print("Updating database with scores...")
    for _, row in df.iterrows():
        conn.execute(
            """UPDATE backtest_results SET composite_score = ?, status = ?
               WHERE strategy_id = ? AND symbol = ? AND timeframe = ?""",
            (row["composite_score"], row["status"],
             row["strategy_id"], row["symbol"], row["timeframe"]),
        )

    # Per-strategy best classification
    best_per_strat = df.groupby("strategy_id").agg({
        "composite_score": "max",
        "win_rate": "max",
        "total_return": "max",
        "status": lambda x: "GEM" if "GEM" in x.values else ("PASS" if "PASS" in x.values else "TRASH"),
    }).reset_index()

    for _, row in best_per_strat.iterrows():
        update_strategy_status(conn, row["strategy_id"], row["status"].lower())

    conn.commit()

    # ── Ticker name mapping for folder names ────
    TICKER_FOLDERS = {
        "BTCUSDT": "btc",
        "ETHUSDT": "eth",
        "SOLUSDT": "sol",
        "^NDX":    "nas100",
        "^DJI":    "us30",
        "^GSPC":   "spx",
        "NDX":     "nas100",
        "DJI":     "us30",
        "GSPC":    "spx",
    }

    # Move GEM files into ticker-specific folders
    # For each GEM backtest result, copy the strategy file to best/<ticker>/
    gem_results = df[df["status"] == "GEM"]
    trash_results = best_per_strat[best_per_strat["status"] == "TRASH"]

    gem_count = 0
    trash_count = 0
    ticker_counts = {}

    for _, row in gem_results.iterrows():
        strat = conn.execute(
            "SELECT source_file FROM strategies WHERE id = ?",
            (row["strategy_id"],),
        ).fetchone()
        if not strat or not strat["source_file"]:
            continue
        src = Path(strat["source_file"])
        if not src.exists():
            continue

        # Determine ticker folder
        symbol = row["symbol"]
        ticker_folder = TICKER_FOLDERS.get(symbol, symbol.lower().replace("^", ""))
        dest_dir = BEST_DIR / ticker_folder
        dest_dir.mkdir(parents=True, exist_ok=True)

        dest = dest_dir / src.name
        if not dest.exists():
            shutil.copy2(src, dest)
            gem_count += 1

        ticker_counts[ticker_folder] = ticker_counts.get(ticker_folder, 0) + 1

    for _, row in trash_results.iterrows():
        strat = conn.execute(
            "SELECT source_file FROM strategies WHERE id = ?",
            (row["strategy_id"],),
        ).fetchone()
        if strat and strat["source_file"]:
            src = Path(strat["source_file"])
            if src.exists():
                dest = ARCHIVE_DIR / src.name
                if not dest.exists():
                    shutil.copy2(src, dest)
                trash_count += 1

    conn.close()

    # Summary
    status_counts = df["status"].value_counts()
    print(f"\n{'=' * 60}")
    print("SCORING COMPLETE")
    print(f"  Total results scored: {len(df):,}")
    print(f"  GEM:   {status_counts.get('GEM', 0):,}")
    print(f"  PASS:  {status_counts.get('PASS', 0):,}")
    print(f"  TRASH: {status_counts.get('TRASH', 0):,}")
    print(f"  Files moved to best/:    {gem_count}")
    print(f"  Files moved to archive/: {trash_count}")
    if ticker_counts:
        print(f"\n  GEMs by ticker:")
        for ticker, count in sorted(ticker_counts.items(), key=lambda x: -x[1]):
            print(f"    strategies/best/{ticker}/  → {count} GEMs")
    print(f"{'=' * 60}")

    # Generate Excel report
    if generate_report:
        report_path = generate_excel_report(df)
        print(f"\n  Excel report: {report_path}")

    return df


def generate_excel_report(df):
    """Generate a color-coded Excel ranking report with per-ticker sheets."""
    report_dir = ROOT / "monitor" / "reports"
    report_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = report_dir / f"strategy_rankings_{ts}.xlsx"

    TICKER_NAMES = {
        "BTCUSDT": "BTC", "ETHUSDT": "ETH", "SOLUSDT": "SOL",
        "^NDX": "NAS100", "^DJI": "US30", "^GSPC": "SPX",
        "NDX": "NAS100", "DJI": "US30", "GSPC": "SPX",
    }

    cols = [
        "strategy_name", "symbol", "timeframe",
        "total_trades", "win_rate", "total_return", "max_drawdown",
        "profit_factor", "sharpe_ratio", "composite_score", "status",
    ]

    gem_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    pass_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    trash_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    wb = Workbook()

    def write_sheet(ws, data, title):
        ws.title = title
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for ri, (_, row) in enumerate(data.head(500).iterrows(), 2):
            status = row.get("status", "PASS")
            fill = gem_fill if status == "GEM" else trash_fill if status == "TRASH" else pass_fill
            for ci, col in enumerate(cols, 1):
                val = row.get(col, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = fill
        for ci in range(1, len(cols) + 1):
            ws.column_dimensions[chr(64 + min(ci, 26))].width = 18

    # Sheet 1: All results (top 500)
    write_sheet(wb.active, df, "All Rankings")

    # Sheet 2: GEMs only
    gems_df = df[df["status"] == "GEM"]
    if not gems_df.empty:
        write_sheet(wb.create_sheet(), gems_df, "ALL GEMS")

    # Per-ticker sheets (GEMs only)
    for symbol in df["symbol"].unique():
        ticker_gems = df[(df["symbol"] == symbol) & (df["status"] == "GEM")]
        if ticker_gems.empty:
            continue
        sheet_name = TICKER_NAMES.get(symbol, symbol.replace("^", ""))
        write_sheet(wb.create_sheet(), ticker_gems, sheet_name)

    wb.save(path)
    return path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="GB-BRAIN Score & Sort")
    parser.add_argument("--report", "-r", action="store_true",
                        help="Generate Excel report")
    args = parser.parse_args()

    run_scoring(generate_report=args.report)
